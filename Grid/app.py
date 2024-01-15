#!/usr/bin/env python
# coding: utf-8

# In[232]:


import skimage.io as io
import matplotlib.pyplot as plt
import numpy as np
from skimage.exposure import histogram
from matplotlib.pyplot import bar
from skimage.color import rgb2gray,rgb2hsv
from pathlib import Path
# Convolution:
from scipy.signal import convolve2d
from scipy import fftpack
import math
from functools import cmp_to_key
from skimage.util import random_noise
from skimage.filters import median
from skimage.feature import canny
import commonfunctions as cf # this a custom module found the commonfunctions.py
import cv2
import numpy as np
import os


# Edges
from skimage.filters import sobel_h, sobel, sobel_v,roberts, prewitt

# Show the figures / plots inside the notebook
def show_images(images,titles=None):
    #This function is used to show image(s) with titles by sending an array of images and an array of associated titles.
    # images[0] will be drawn with the title titles[0] if exists
    # You aren't required to understand this function, use it as-is.
    n_ims = len(images)
    if titles is None: titles = ['(%d)' % i for i in range(1,n_ims + 1)]
    fig = plt.figure()
    n = 1
    for image,title in zip(images,titles):
        a = fig.add_subplot(1,n_ims,n)
        if image.ndim == 2: 
            plt.gray()
        plt.imshow(image)
        a.set_title(title)
        n += 1
    fig.set_size_inches(np.array(fig.get_size_inches()) * n_ims)
    plt.show() 


def showHist(img):
    # An "interface" to matplotlib.axes.Axes.hist() method
    plt.figure()
    imgHist = histogram(img, nbins=256)
    
    bar(imgHist[1].astype(np.uint8), imgHist[0], width=0.8, align='center')


alpha = 1.0
beta = 0.0
# Function to perform the perspective transformation
def perspective_transform(img,binary):

    if img is None:
        print(f"Error: Unable to load the image.")
    else:
        # Convert the image to grayscale
        binary_image=0
        if binary==0:
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

            # Apply thresholding to create a binary image
            _, binary_image = cv2.threshold(gray, 50, 255, cv2.THRESH_BINARY)
        else :
            binary_image=img
        # cf.show_images([binary_image])
        # Find contours in the binary image
        contours, _ = cv2.findContours(binary_image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        # Find the largest contour based on area
        print(len(contours))
        largest_contour = max(contours, key=cv2.contourArea)
        # Approximate the contour to a polygon
        epsilon = 0.02 * cv2.arcLength(largest_contour, True)
        approx_polygon = cv2.approxPolyDP(largest_contour, epsilon, True)

        # Get the four corners of the polygon
        corners = approx_polygon.reshape(-1, 2)
        corners = sorted(corners, key=lambda x: x[1])
        # Separate the sorted corners into top and bottom
        top_corners = sorted(corners[:2], key=lambda x: x[0])
        bottom_corners = sorted(corners[2:], key=lambda x: x[0])

        # Concatenate the sorted corners
        sorted_corners = np.concatenate([bottom_corners, top_corners])

        # Define the destination points for the perspective transformation
        dst_points = np.float32([[0, img.shape[0]], [img.shape[1], img.shape[0]], [0, 0], [img.shape[1], 0]])

        # Calculate the perspective transformation matrix
        matrix = cv2.getPerspectiveTransform(sorted_corners.astype(np.float32), dst_points)

        # Apply the perspective transformation to the image
        warped_img = cv2.warpPerspective(img, matrix, (img.shape[1], img.shape[0]))
        return warped_img
def invert_image(img):
    clone =img
    adjusted_img = cv2.addWeighted(img, alpha, np.zeros(img.shape, img.dtype), 0, beta)
    # cf.show_images([adjusted_img])
    img=perspective_transform(adjusted_img,0)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    thresh_image = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY,91, 5)
    thresh_image=255-thresh_image
    trial = perspective_transform(thresh_image,1)
    # cf.show_images([thresh_image])
    transform=perspective_transform(thresh_image,1)
    return transform


# In[ ]:





# In[233]:



arr=[[] for _ in range (6)]
def readImage(imgpath):
    img=cv2.imread(imgpath)
    transformed_img = perspective_transform(img, 0)
    inverted_img=invert_image(transformed_img)
    show_images([inverted_img])
    return inverted_img
def getVerticalLines(inverted_img):
    # Apply edge detection using Canny edge detector
    edges = cv2.Canny(inverted_img, 50, 150, apertureSize=3)
    # Perform Hough Line Transform for vertical lines
    rho = 1  # 1 pixel
    theta = np.pi # Vertical lines (90 degrees)
    threshold = 260  # Adjust this value as needed
    min_line_length = 160  # Minimum line length
    max_line_gap = 20  # Maximum allowed gap between line segments

    lines = cv2.HoughLinesP(edges, rho=rho, theta=theta, threshold=threshold,
                            minLineLength=min_line_length, maxLineGap=max_line_gap)
    sorted_lines = sorted(lines, key=lambda line: line[0][0])
    print("from vertical lines",sorted_lines)
    filtered_lines = [sorted_lines[0]]
    for line in sorted_lines[1:]:
        prev_x = filtered_lines[-1][0][0] 
        cur_x = line[0][0]  

        if cur_x - prev_x >= 20:
            filtered_lines.append(line)  
    return filtered_lines
def getBlocks(lines,cur_Seg,inverted_img):
    x1=lines[0][0][0]
    idx=0
    for line in lines:
        x2=line[0][0]
        if(x2-x1>cur_Seg[0]):
            block = inverted_img[0:4000, x1:x2]
            break      
        idx=idx+1
    y1 = 150
    while y1 < block.shape[0]:
        cell = block[y1:y1 + 200, :]
        show_images([cell])
        arr[cur_Seg[1]].append(cell)
        y1 += 220
    return [block,idx]
segs=[[50,0],[100,1],[200,2],[50,3],[50,4],[50,5]]
def getImageBlocks(filtered_lines,inverted_img):
    idx = 0
    block_images = [] 
    for i in range(len(segs)):
        [block, nwidx] = getBlocks(filtered_lines[idx:], segs[i],inverted_img)
        idx = idx + nwidx
        block_images.append(block)  
    return block_images
    
    
samples_dir = './Samples/'
for i in range(8, 9):
    print(i)
    image_path = os.path.join(samples_dir, f"{i}.jpg")
    inverted_img = readImage(image_path)
    filtered_lines = getVerticalLines(inverted_img)
    blocks=getImageBlocks(filtered_lines, inverted_img)
    # for j in range (0,1):
    #     getcells(blocks[j])
    block_titles = [f"Image: {i} - Block {j+1}" for j in range(len(blocks))]
    show_images(blocks, block_titles)


# In[ ]:





# In[234]:


def contour_sort(a, b):

    br_a = cv2.boundingRect(a)
    br_b = cv2.boundingRect(b)

    if abs(br_a[0] - br_b[0]) <= 5:
        return br_a[1] - br_b[1]

    return br_a[0] - br_b[0]


def sorted_counter (contours):

    return sorted(contours, key=cmp_to_key(contour_sort))


# In[235]:



def largest_contour(contours):
    largest_contour = np.array([])
    max_area = 0
    for contour in contours:
        area = cv2.contourArea(contour)
        perimeter = cv2.arcLength(contour, True)
        approx = cv2.approxPolyDP(contour, 0.1 * perimeter, True)
        if area > max_area and len(approx) == 4:
            largest_contour = approx
            max_area = area
    return largest_contour
def Prespective_Transform(img_original):
    img =img_original.copy()
    gray_image = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    gray_image = cv2.bilateralFilter(gray_image, 20, 30, 30)

    edges = cv2.Canny(gray_image, 20, 120)
    edges = cv2.dilate(edges.copy(), None, 2)

    contours, hierarchy = cv2.findContours(edges, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

    contours = sorted(contours, key=cv2.contourArea, reverse=True)

    biggest = largest_contour(contours[2:10])
    if(biggest.size!=8):
        biggest = largest_contour(contours[0:1])
    
    pts = biggest.reshape(4, 2)

    pts_sum = pts.sum(axis=1)

    top_left = pts[np.argmin(pts_sum)]
    bottom_right = pts[np.argmax(pts_sum)]

    pts_diff = np.diff(pts, axis=1)
    top_right = pts[np.argmin(pts_diff)]
    bottom_left = pts[np.argmax(pts_diff)]


    #Image Dimensions
    bottom_width = np.sqrt(pow((bottom_right[0] - bottom_left[0]), 2) + (pow((bottom_right[1] - bottom_left[1]), 2)))
    top_width = np.sqrt(pow((top_right[0] - top_left[0]), 2) + (pow((top_right[1] - top_left[1]), 2)))
    right_height = np.sqrt(pow((top_right[0] - bottom_right[0]), 2) + (pow((top_right[1] - bottom_right[1]), 2)))
    left_height = np.sqrt(pow((top_left[0] - bottom_left[0]), 2) + (pow((top_left[1] - bottom_left[1]), 2)))


    # Output image size
    width = max(int(bottom_width), int(top_width))
    height = max(int(right_height), int(left_height))
    # Points with new Coordinates 
    converted_points = np.float32([[0, 0], [width, 0], [0, height], [width, height]])
    arr = np.float32([top_left, top_right, bottom_left, bottom_right])
    matrix = cv2.getPerspectiveTransform(arr, converted_points)
    img_output = cv2.warpPerspective(img_original, matrix, (width, height))
    return img_output


# In[236]:



def saveImg(directory, suffix, id,img):
    Path(directory).mkdir(parents=True, exist_ok=True)
    cv2.imwrite(f"{directory}/{suffix}_{id}.jpg", img)
def getCells(img):
    img_bin = cv2.threshold(img, 128, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    (thresh, img_bin) = cv2.threshold(img, 128, 255,cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    img_bin = 255 - img_bin

    filter_dim = np.array(img).shape[1] // 38
    
    v_filter = cv2.getStructuringElement(cv2.MORPH_RECT, (1, filter_dim))
    h_filter = cv2.getStructuringElement(cv2.MORPH_RECT, (filter_dim, 1))
    filter = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))

    v_lines = cv2.erode(img_bin, v_filter, iterations=3)
    v_lines = cv2.dilate(v_lines, v_filter, iterations=3)
    lines_ver = cv2.HoughLinesP(v_lines, 1, np.pi / 180, 40, minLineLength=10, maxLineGap=20)

    for line in lines_ver:
        for x1, y1, x2, y2 in line:
            v_lines = cv2.line(v_lines, (x1, 0), (x2, v_lines.shape[0]), (255, 255, 255), 1)

    h_lines = cv2.erode(img_bin, h_filter, iterations=4)
    h_lines = cv2.dilate(h_lines, h_filter, iterations=3)
    lines_hor = cv2.HoughLinesP(h_lines, 2, np.pi / 180, 40, minLineLength=5, maxLineGap=10)

    for line in lines_hor:
        for x1, y1, x2, y2 in line:
            h_lines = cv2.line(h_lines, (0, y1), (h_lines.shape[1], y2), (255, 255, 255), 1)

    final_img = cv2.bitwise_and(v_lines, h_lines)
    final_img = cv2.erode(~final_img, filter, iterations=1)
    (thresh, img_output) = cv2.threshold(final_img, 128,255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU) 
    show_images([img_output])
    return img_output


def cutCells(img, orignal_img,output_dir):
    contours = cv2.findContours(img, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)[0]
    contours = sorted_counter(contours)
    rows = []

    for c in range(len(contours)-1):
        x1, y1, w1, h1 = cv2.boundingRect(contours[c])
        x2, y2, w2, h2 = cv2.boundingRect(contours[c+1])
        
        if x1 == x2:
            rows.append(y1)
        else:
            rows.append(y1)
            break

    num_hor = len(rows)
    num_ver = len(contours) 
    arr=["Code","Student Name","English Name","1","2","3"]
    for col in range(0, 6):
        for row in range(1, num_hor-1):
            x1, y1, w1, h1 = cv2.boundingRect(contours[row+num_hor*col])
            x2, y2, w2, h2 = cv2.boundingRect(contours[row+1+num_hor*col])
            x3, y3, w3, h3 = cv2.boundingRect(contours[row+num_hor*col+num_hor+1])
            new_img = orignal_img[y1+h1:y3, x2+w2:x3]
            # show_images([new_img])
            saveImg(output_dir,arr[col],row,new_img)


# In[237]:


directory = './Samples'
for filename in os.listdir(directory):
        filepath = os.path.join(directory, filename)
        print(filename)
        img = cv2.imread(filepath)
        cur_image = Prespective_Transform(img)
        gray_image = cv2.cvtColor(cur_image, cv2.COLOR_BGR2GRAY)
        img_final = getCells(gray_image)
        image_name = os.path.splitext(filename)[0]
        output_dir = f"results/{image_name}"
        cutCells(img_final, gray_image,output_dir)


# In[ ]:





# In[238]:


# # Install the required libraries
# pip install pytesseract pillow

##author:Mohamed Samir
## this cell is defined for all detection algorithms function to be used
#make sure to install all dependencies
import cv2
import pytesseract
from PIL import Image

# pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

#this does work on some hand written and doesnt work on else
def ocr_pytesseract_number_extraction(image):
    # Open the image using Pillow
    #you can remove the config to detect the text if you want but we only using it for digits detection
    extracted_text = pytesseract.image_to_string(image, config='--psm 6 -c tessedit_char_whitelist=0123456789')
    return extracted_text
print(ocr_pytesseract_number_extraction('./results/1/1_9.jpg'))


# In[239]:


import os
class Sample:
    def __init__(self, studentCode, q1, q2,q3):
        self.studentCode = studentCode
        self.q1 = q1
        self.q2 = q2
        self.q3=q3
    def display(self):
        print(f"Student Code: {self.studentCode}\n Q1: {self.q1}\n Q2: {self.q2}")


def get_countOfFolder(folder_path):
    try:
        # List all items (files and subdirectories) in the folder
        items = os.listdir(folder_path)

        # Count the number of items (excluding hidden files and directories)
        num_items = len([item for item in items if not item.startswith('.')])

        return num_items
    except Exception as e:
        print(f"Error counting items in {folder_path}: {e}")
        return -1

def sorting_keyResults(file_name):
        try:
            return int(file_name.split('_')[1].split('.')[0])
        except (IndexError, ValueError):
            return float('inf')  

def sorting_key_resultFolders(file_name):
        try:
            return int(file_name)
        except (IndexError, ValueError):
            return float('inf')  

def get_files(directory_path):
    files = os.listdir(directory_path)
    return files

def get_resultFiles_sorted(directory_path,startWord):
    # print(directory_path.split('/')[1].split('s')[0])
    files = os.listdir(directory_path)
    code_files = [file for file in files if file.startswith(startWord)]
    print(code_files)
    code_files.sort(key=sorting_keyResults)
    return code_files

#first photo 
Question1='./results/1/'




# In[240]:


import os
directory_path='./results/'
result_subDirectories=get_files(directory_path)
result_subDirectories.sort(key=sorting_key_resultFolders)
print(result_subDirectories)

# for subdirectory in result_subDirectories:
    






# In[241]:


#this cell is for loading the data set as a hog features and then labeling each feature
#after that it saves the model as a .joblib file

##NOTEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEE##
##You dont need to run this cell as the model file is already provided##

from sklearn.neighbors import KNeighborsClassifier
from sklearn.neural_network import MLPClassifier
from sklearn.model_selection import train_test_split
from sklearn import svm
import numpy as np
import os
import cv2
from joblib import dump, load

# Define the random seed
random_seed = 42  # You can use any integer value

target_img_size = (32, 32)
classifiers = {
    'SVM': svm.LinearSVC(random_state=random_seed),
}

def extract_hog_features(img):
    """
    TODO
    You won't implement anything in this function. You just need to understand it 
    and understand its parameters (i.e win_size, cell_size, ... etc)
    """
    img = cv2.resize(img, dsize=target_img_size)
    win_size = (32, 32)
    cell_size = (4, 4)
    block_size_in_cells = (2, 2)
    
    block_size = (block_size_in_cells[1] * cell_size[1], block_size_in_cells[0] * cell_size[0])
    block_stride = (cell_size[1], cell_size[0])
    nbins = 9  # Number of orientation bins
    hog = cv2.HOGDescriptor(win_size, block_size, block_stride, cell_size, nbins)
    h = hog.compute(img)
    h = h.flatten()
    return h.flatten()

# Function to load dataset images
def load_dataset_digits(root_folder):
    features = []
    labels = []
    for digit in range(10):
        symbol_folder = os.path.join(root_folder, str(digit))
        symbol_files = os.listdir(symbol_folder)
        for filename in symbol_files:
            if filename.endswith(".jpg"):
                img_path = os.path.join(symbol_folder, filename)
                img = cv2.imread(img_path, 0)  # Load image in grayscale
                if img is not None:
                    features.append(extract_hog_features(img))
                    labels.append(digit)
    return np.array(features), np.array(labels)

features_digits, labels_digits = load_dataset_digits('./dataset/digits/')
digits_trainFeatures, digits_testFeatures, digits_trainLabels, digits_testLabels = train_test_split(
        features_digits, labels_digits, test_size=0.2, random_state=random_seed)


digits_models = {}
for model_name, model in classifiers.items():
    print('############## Training digits_dataset', model_name, "##############")
    
    # Train the model only on the training features
    model.fit(digits_trainFeatures, digits_trainLabels)
    
    # Test the model on images it hasn't seen before
    accuracy = model.score(digits_testFeatures, digits_testLabels)
    
    print(model_name, 'accuracy:', accuracy * 100, '%')

    digits_models[model_name] = model

dump(digits_models, 'digits_models.joblib')
print('Saved digits models to digits_models.joblib')
    # Function to load dataset symbols
def load_dataset_symbols(root_folder):
    features = []
    labels = []
    symbols=os.listdir(root_folder)
    # print(symbols)
    for symbol in symbols:
        symbol_folder = os.path.join(root_folder + symbol)
        symbol_files = os.listdir(symbol_folder)
        # print(symbol_files)
        for filename in symbol_files:
            if filename.endswith(".jpg"):
                img_path = os.path.join(symbol_folder, filename)
                img = cv2.imread(img_path, 0)  # Load image in grayscale
                if img is not None:
                    x=extract_hog_features(img)
                    features.append(x)
                    # print(features)
                    labels.append(symbol)
    return np.array(features), np.array(labels)



    
features_symbols , labels_symbols =load_dataset_symbols('./dataset/symbols/')
symbols_trainFeatures, symbols_testFeatures,symbols_trainLabels,symbols_testLabels=train_test_split(features_symbols,labels_symbols,test_size=0.2,random_state=random_seed)

symbols_models = {}

for model_name, model in classifiers.items():
    print('############## Training Load_symbols', model_name, "##############")
    
    # Train the model only on the training features
    model.fit(symbols_trainFeatures, symbols_trainLabels)
    
    # Test the model on images it hasn't seen before
    accuracy = model.score(symbols_testFeatures, symbols_testLabels)
    
    print(model_name, 'accuracy:', accuracy * 100, '%')

    symbols_models[model_name] = model

dump(symbols_models, 'symbols_models.joblib')
print('Saved symbols models to symbols_models.joblib')


# In[242]:



def load_model():
    # Load the digits models dictionary
    loaded_digits_models = load('digits_models.joblib')
    print('Loaded digits models from digits_models.joblib')
    print(loaded_digits_models)
    # Load the symbols models dictionary
    loaded_symbols_models = load('symbols_models.joblib')
    print('Loaded symbols models from symbols_models.joblib')


# def predict_symbol(img):

load_model()


def predict_symbol(img):

    test_features=extract_hog_features(img)
    predicted_symbol=loaded_symbols_models['SVM'].predict([test_features])
    return predicted_symbol



def predict_digit(img):
    test_features=extract_hog_features(img)
    predicted_digit=loaded_digits_models['SVM'].predict([test_features])
    return predicted_digit



img=cv2.imread('./results/1/1_2.jpg',0)
show_images([img])
print(predict_digit(img))






# In[263]:


import cv2
import pytesseract
import numpy as np
import os

def code_split(img):
    # Check if the image is already in grayscale
    if len(img.shape) == 2:
        gray = img
    else:
        # Convert the image to grayscale
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # Use adaptive thresholding to create a binary image
    _, thresh = cv2.threshold(gray, 120, 255, cv2.THRESH_BINARY_INV)
    # Experiment with the threshold value (128 in this case) to suit your needs

    # Find contours in the binary image
    contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    contours = sorted(contours, key=lambda x: cv2.boundingRect(x)[0])
    
    # Array to store the extracted digits
    digit_images = []

    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        digit = gray[y:y + h, x:x + w]

        digit_images.append(digit)

    return digit_images



def SVM_getSampleImages(directory_path):
    # loop inside the 
    files = os.listdir(directory_path)
    code_files = [file for file in files if file.startswith('Code')]
    code_files.sort(key=sorting_qColumns)
    
    q1_files = [file for file in files if file.startswith('1')]
    q2_files = [file for file in files if file.startswith('2')]
    q3_files = [file for file in files if file.startswith('3')]
    
    q1_files.sort(key=sorting_qColumns)
    q2_files.sort(key=sorting_qColumns)
    q3_files.sort(key=sorting_qColumns)
    
    code_images, q1_images, q2_images, q3_images = [], [], [], []
    
    for i in range(len(q1_files)):
        code_path = os.path.join(directory_path, code_files[i])
        code_img = cv2.imread(code_path, 0)  # Load image in grayscale
        code_images.append(code_img)
        
        q1_path = os.path.join(directory_path, q1_files[i])
        q1_img = cv2.imread(q1_path, 0)  # Load image in grayscale
        q1_images.append(q1_img)
        
        q2_path = os.path.join(directory_path, q2_files[i])
        q2_img = cv2.imread(q2_path, 0)  # Load image in grayscale
        q2_images.append(q2_img)
        
        q3_path = os.path.join(directory_path, q3_files[i])
        q3_img = cv2.imread(q3_path, 0)  # Load image in grayscale
        q3_images.append(q3_img)
    
    return np.array(code_images), np.array(q1_images), np.array(q2_images), np.array(q3_images)

def code_predict(code_images, key):
    # key-> 'OCR' we choose ocr
    # key-> 'SVM' we choose SVM
    code_list = []

    if key == 'OCR':
        for image in code_images:
            code = ocr_pytesseract_number_extraction(image)
            code_list.append(code)
    elif key == 'SVM':
        for image in code_images:
            digit_images = code_split(image)
            code = ""
            for digit in digit_images:
                predicted = predict_digit(digit)
                # print(predicted)
                code += str(predicted[0])
            # print(code)
            code_list.append(code)
    else:
        print("No Key is defined as " + key)

    return code_list

def q1_predict(q1_images, key):
    # key-> 'OCR' we choose ocr
    # key-> 'SVM' we choose SVM
    q1_list = []

    if key == 'OCR':
        for image in q1_images:
            code = ocr_pytesseract_number_extraction(image)
            q1_list.append(code)
    elif key == 'SVM':
           for image in q1_images:
                code = predict_digit(image)
                q1_list.append(str(code[0]))
    else:
        print("No Key is defined as " + key)

    return q1_list


def q2_predict(q2_images):
    # key-> 'OCR' we choose ocr
    # key-> 'SVM' we choose SVM
    q2_list = []
    for image in q2_images:
        symbol = predict_symbol(image)
        if(str(symbol[0])=="correct"):
            q2_list.append("5")
        elif(str(symbol[0])=="box"):
            q2_list.append("0")   
        elif(str(symbol[0]).startswith("horizontal")):
            # if(str(symbol[0].split('l')[1]=='1')):
            #     q2_list.append('0')
            # else:
            q2_list.append(str(5-int(str(symbol[0]).split('l')[1])))
        elif(str(symbol[0]).startswith("vertical")):
            q2_list.append(str(symbol[0]).split('l')[1])
        else:
            q2_list.append(str(symbol[0]))
        
    return q2_list



# digits->digits
# empty->empty
# box->box
# ?->?
# check->5
# horizontal->5-lines
# vertical->lines



# In[264]:


code_images, q1_images, q2_images, q3_images=SVM_getSampleImages('./results/1/')
code_list=code_predict(code_images,'SVM')
q1_list=q1_predict(q1_images,'SVM')
q2_list=q2_predict(q2_images)
q3_list=q2_predict(q3_images)
# print(code_list)
for i in range( len(code_list)):
    code_list[i]=code_list[i].strip()
    q1_list[i]=q1_list[i].strip()
print(code_list)
print(q1_list)
show_images(q2_images)
print(q2_list)
print(q3_list)


# In[287]:


import openpyxl
from openpyxl.styles import PatternFill

def open_excel(ids,data_col1,data_col2,data_col3,i):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Set headers for columns
    sheet['A1'] = 'Code'
    sheet['B1'] = '1'
    sheet['C1'] = '2'
    sheet['D1'] = '3'

    # Fill the columns with corresponding data and set specific cells to have a red background color
    for i, id_value in enumerate(ids, start=2):  # Start from row 2 (as 1st row contains headers)
        sheet.cell(row=i, column=1, value=id_value)  # Fill IDs in column A
        
        sheet.cell(row=i, column=2, value=data_col1[i - 2])
    
        
        
        if data_col2[i-2] == 'question':  
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            sheet.cell(row=i, column=3).fill = red_fill  # Set column 3 cell to red
        elif data_col2[i-2] != 'empty':
            sheet.cell(row=i, column=3, value=data_col2[i - 2])
        if   data_col3[i-2] == 'question':  
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            sheet.cell(row=i, column=4).fill = red_fill  
        elif data_col3[i-2] != 'empty':
            sheet.cell(row=i, column=4, value=data_col3[i - 2])

    # Save the workbook
    workbook.save(f'res{i}.xlsx')


# In[288]:


import os
# code_images, q1_images, q2_images, q3_images=SVM_getSampleImages('./results/1/')
# code_list=code_predict(code_images,'SVM')
# q1_list=q1_predict(q1_images,'SVM')
# q2_list=q2_predict(q2_images)
# q3_list=q2_predict(q3_images)
# # print(code_list)
# for i in range( len(code_list)):
#     code_list[i]=code_list[i].strip()
#     q1_list[i]=q1_list[i].strip()
# print(code_list)
# print(q1_list)
# show_images(q2_images)
# print(q2_list)
# print(q3_list)

def get_samples_data(directory_path,algorithm):
    check='SVM'
    if(algorithm==1):
        check='OCR'
    elif(algorithm==0):
        check='SVM'
    folders=os.listdir(directory_path)
    for folder in folders:
        folder_path=os.path.join(directory_path, folder + '/')
        print(folder_path)
        code_images, q1_images, q2_images, q3_images=SVM_getSampleImages(folder_path)
        code_list=code_predict(code_images,check)
        q1_list=q1_predict(q1_images,check)
        q2_list=q2_predict(q2_images)
        q3_list=q2_predict(q3_images)
        # print(code_list)
        for i in range( len(code_list)):
            code_list[i]=code_list[i].strip()
            q1_list[i]=q1_list[i].strip()
        open_excel(code_list,q1_list,q2_list,q3_list,int(folder))
    # print(folders)


get_samples_data('./results/',0)

